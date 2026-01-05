import pandas as pd
import ipaddress
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ============== CONFIGURATION ==============
# Change these values as needed
HOSTNAME_PREFIX = "cam"  # Change to 'nvr', 'sw', etc.
STARTING_NUMBER = 1      # Starting number for hostnames
DEVICE_TYPE = "Camera"   # Change to 'Switch', 'Firewall', 'Camera', etc.
NVR_NUMBER = 1           # The 'x' in NVRx
CAM_STARTING_NUMBER = 1  # The starting 'y' in camy
INPUT_FILE = "input_ips.xlsx"  # Your input Excel file
OUTPUT_FILE = "nvr1.xlsx"  # Output file name

# Available subnets
AVAILABLE_SUBNETS = [
    "10.1.0.0/22",
    "10.10.0.0/21",
    "10.10.16.0/23",
    "10.10.19.0/24",
    "10.10.20.0/24",
    "10.10.21.0/24",
    "10.10.22.0/24",
    "10.10.23.0/24",
    "10.10.24.0/24",
    "10.10.25.0/24",
    "10.10.26.0/24",
    "10.10.27.0/24",
    "10.10.28.0/24",
    "10.10.29.0/24",
    "10.10.30.0/24",
    "10.10.31.0/24",
    "10.10.32.0/24",
    "10.10.33.0/24",
    "10.10.34.0/24",
    "10.10.35.0/24",
    "10.10.36.0/24",
    "10.10.39.0/24"
]
# ============================================

def find_matching_subnet(ip_address, subnets):
    """Find the subnet that contains the given IP address"""
    try:
        ip = ipaddress.ip_address(ip_address)
        for subnet_str in subnets:
            subnet = ipaddress.ip_network(subnet_str, strict=False)
            if ip in subnet:
                return str(subnet)
        return ""  # Return empty if no match found
    except:
        return ""

def create_formatted_excel(input_file, output_file, hostname_prefix, starting_num, 
                          subnets, device_type, nvr_num, cam_start_num):
    """Create formatted Excel file from input data"""
    
    # Read input Excel file
    # Assuming input has columns: 'ip' and 'location'
    df_input = pd.read_excel(input_file)
    
    # Create output dataframe with all required columns
    columns = [
        'section *', 'ip address *', 'hostname', 'description', 'vrf',
        'subnet *', 'mac', 'owner', 'device', 'note', 'tag',
        'is_gateway', 'custom_device_type', 'custom_Location', 'custom_Hospital-Location'
    ]
    
    # Initialize empty dataframe
    df_output = pd.DataFrame(columns=columns)
    
    # Process each IP address
    for idx, row in df_input.iterrows():
        ip = row['ip']
        location = row.get('location', '')
        
        # Generate hostname
        hostname = f"{hostname_prefix}{starting_num + idx:02d}"
        
        # Find matching subnet
        subnet = find_matching_subnet(ip, subnets)
        
        # Generate custom_Location (e.g., nvr2 cam1, nvr2 cam2, etc.)
        custom_location = f"nvr{nvr_num} cam{cam_start_num + idx}"
        
        # Create new row
        new_row = {
            'section *': 'customers',
            'ip address *': ip,
            'hostname': hostname,
            'description': '',
            'vrf': '',
            'subnet *': subnet,
            'mac': '',
            'owner': '',
            'device': '',
            'note': '',
            'tag': '',
            'is_gateway': '',
            'custom_device_type': device_type,
            'custom_Location': custom_location,
            'custom_Hospital-Location': location
        }
        
        df_output = pd.concat([df_output, pd.DataFrame([new_row])], ignore_index=True)
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Network Devices"
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_num, row_data in enumerate(df_output.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
    
    # Adjust column widths
    column_widths = {
        'A': 15, 'B': 15, 'C': 15, 'D': 20, 'E': 12,
        'F': 18, 'G': 18, 'H': 12, 'I': 15, 'J': 12,
        'K': 12, 'L': 12, 'M': 20, 'N': 20, 'O': 25
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file created successfully: {output_file}")
    print(f"Total devices processed: {len(df_output)}")

# Main execution
if __name__ == "__main__":
    try:
        create_formatted_excel(
            INPUT_FILE, 
            OUTPUT_FILE, 
            HOSTNAME_PREFIX, 
            STARTING_NUMBER, 
            AVAILABLE_SUBNETS,
            DEVICE_TYPE,
            NVR_NUMBER,
            CAM_STARTING_NUMBER
        )
    except FileNotFoundError:
        print(f"Error: Input file '{INPUT_FILE}' not found!")
        print("Please make sure your input Excel file exists and has columns: 'ip' and 'location'")
    except Exception as e:
        print(f"Error: {e}")